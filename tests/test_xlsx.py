import json
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from audit_compiler.adapters.xlsx import parse_xlsx_workbook
from audit_compiler.cli import main
from audit_compiler.duckdb_store import connect, store_xlsx_workbook


def _add_cached_formula(path: Path, *, archive_sheet: str, formula: str, value: str) -> None:
    with ZipFile(path) as source:
        entries = {name: source.read(name) for name in source.namelist()}
    cached_xml = f"<f>{formula}</f><v>{value}</v>".encode()
    candidates = [
        f"<f>{formula}</f><v></v>".encode(),
        f"<f>{formula}</f><v />".encode(),
    ]
    formula_xml = next(
        (candidate for candidate in candidates if candidate in entries[archive_sheet]), None
    )
    assert formula_xml is not None
    entries[archive_sheet] = entries[archive_sheet].replace(formula_xml, cached_xml)
    with ZipFile(path, "w", ZIP_DEFLATED) as target:
        for name, content in entries.items():
            target.writestr(name, content)


def _write_multisheet_workbook(path: Path) -> None:
    workbook = Workbook()
    ledger = workbook.active
    ledger.title = "Ledger"
    ledger["A1"] = "Quarterly export"
    headers = ["Reference", "Amount", "Booking date", "Optional", "Calculated"]
    for column, header in enumerate(headers, start=1):
        ledger.cell(row=3, column=column, value=header)
    ledger["A4"] = "R-1"
    ledger["B4"] = "1.234,50"
    ledger["C4"] = date(2026, 7, 18)
    ledger["E4"] = "=1+1"
    ledger["A5"] = "R-2"
    ledger["B5"] = "12,00"
    ledger["C5"] = "19.07.2026"

    notes = workbook.create_sheet("Notes")
    notes.append(["Key", "Value"])
    notes.append(["Language", "Deutsch"])
    workbook.save(path)
    _add_cached_formula(
        path,
        archive_sheet="xl/worksheets/sheet1.xml",
        formula="1+1",
        value="2",
    )


def test_xlsx_preserves_cells_metadata_normalization_and_cached_formula(tmp_path: Path) -> None:
    path = tmp_path / "workbook.xlsx"
    _write_multisheet_workbook(path)

    workbook = parse_xlsx_workbook(path, dossier_root=tmp_path)

    assert [sheet.name for sheet in workbook.sheets] == ["Ledger", "Notes"]
    ledger = workbook.sheets[0]
    assert ledger.dimension == "A1:E5"
    assert ledger.header_row == 3
    assert ledger.metadata_row_numbers == (1, 2)
    assert len(ledger.records) == 2

    cells = {cell.coordinate: cell for cell in ledger.cells}
    assert cells["A1"].raw_value == "Quarterly export"
    assert cells["A1"].evidence_ref.sheet == "Ledger"
    assert cells["A1"].evidence_ref.cell == "A1"
    assert cells["B4"].raw_value == "1.234,50"
    assert cells["B4"].normalized_value == "1234.50"
    assert cells["B4"].normalization_kind == "decimal"
    assert cells["C4"].normalized_value == "2026-07-18"
    assert cells["D4"].raw_value == ""
    assert cells["D4"].normalized_value is None
    assert cells["E4"].raw_value == "2"
    assert cells["E4"].formula == "1+1"
    assert cells["E4"].extraction_method.endswith("formula_cache")


def test_xlsx_missing_and_duplicate_headers_are_warned_and_disambiguated(tmp_path: Path) -> None:
    path = tmp_path / "headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Code", None, "Code"])
    sheet.append(["A", "middle", "B"])
    workbook.save(path)

    parsed = parse_xlsx_workbook(path, dossier_root=tmp_path).sheets[0]

    assert parsed.header_row == 1
    assert parsed.headers == ("Code", "column_B", "Code__2")
    messages = [warning.message for warning in parsed.warnings]
    assert any("missing header" in message for message in messages)
    assert any("duplicate header" in message for message in messages)


def test_xlsx_without_header_retains_all_cells_as_metadata_and_warns(tmp_path: Path) -> None:
    path = tmp_path / "no-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([1, 2])
    sheet.append([3, 4])
    workbook.save(path)

    parsed = parse_xlsx_workbook(path, dossier_root=tmp_path).sheets[0]

    assert parsed.header_row is None
    assert parsed.metadata_row_numbers == (1, 2)
    assert parsed.records == ()
    assert len(parsed.cells) == 4
    assert parsed.warnings[0].message == "no conservative header row detected"


def test_xlsx_is_stored_and_reported_by_compile_cli(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    _write_multisheet_workbook(path)
    parsed = parse_xlsx_workbook(path, dossier_root=tmp_path)
    database = tmp_path / "xlsx.duckdb"
    connection = connect(database)

    assert store_xlsx_workbook(connection, parsed) == 3
    assert connection.execute("SELECT count(*) FROM xlsx_workbooks").fetchone()[0] == 1
    assert connection.execute("SELECT count(*) FROM xlsx_sheets").fetchone()[0] == 2
    assert connection.execute("SELECT count(*) FROM xlsx_cells").fetchone()[0] == 29
    assert connection.execute("SELECT count(*) FROM xlsx_metadata_rows").fetchone()[0] == 2
    evidence = connection.execute(
        """
        SELECT source_path, sheet, cell, raw_value, normalized_value, extraction_method
        FROM evidence_refs
        WHERE sheet = 'Ledger' AND cell = 'B4'
        """
    ).fetchone()
    assert evidence == (
        "source.xlsx",
        "Ledger",
        "B4",
        "1.234,50",
        "1234.50",
        "openpyxl.read_only.data_only",
    )
    connection.close()

    report_path = tmp_path / "report.json"
    main(
        [
            "store",
            str(tmp_path),
            "--database",
            str(database),
            "--output",
            str(report_path),
        ]
    )
    report = json.loads(report_path.read_text(encoding="utf-8"))
    workbook_report = next(
        source for source in report["sources"] if source["source_path"] == "source.xlsx"
    )
    assert workbook_report["parsing_status"] == "success"
    assert workbook_report["parsed_row_count"] == 3
    assert [sheet["sheet_name"] for sheet in workbook_report["sheets"]] == [
        "Ledger",
        "Notes",
    ]
    assert workbook_report["sheets"][0]["detected_header_row"] == 3
    assert workbook_report["sheets"][0]["dimension"] == "A1:E5"
