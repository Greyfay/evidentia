"""Native XLSX extraction with cell-level provenance and conservative headers."""

from __future__ import annotations

import posixpath
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Annotated, Any
from uuid import UUID, uuid4
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_FORMULA
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import Field

from audit_compiler.inventory import SourceFile, infer_file_type, sha256_file
from audit_compiler.models import EvidenceRef, ImmutableModel, SourceType
from audit_compiler.normalization import normalize_identifier, parse_date, parse_decimal

_GERMAN_DECIMAL = re.compile(
    r"^\s*(?:EUR|€)?\s*\(?[+-]?(?:\d{1,3}(?:\.\d{3})+|\d+),\d+\)?\s*(?:EUR|€)?\s*$",
    re.IGNORECASE,
)
_GERMAN_DATE = re.compile(r"^\s*\d{1,2}\.\d{1,2}\.\d{2,4}\s*$")

Cells = Annotated[tuple["XlsxCell", ...], Field(min_length=1)]


class XlsxWarning(ImmutableModel):
    message: str = Field(min_length=1)
    sheet_name: str
    row_number: int | None = Field(default=None, ge=1)
    cell: str | None = None


class XlsxCell(ImmutableModel):
    cell_id: UUID = Field(default_factory=uuid4)
    coordinate: str = Field(min_length=2)
    row_number: int = Field(ge=1)
    column_number: int = Field(ge=1)
    raw_value: str
    value_type: str = Field(min_length=1)
    normalized_value: str | None = None
    normalization_kind: str | None = None
    extraction_method: str = Field(min_length=1)
    formula: str | None = None
    evidence_ref: EvidenceRef


class XlsxNormalizedRecord(ImmutableModel):
    record_id: UUID = Field(default_factory=uuid4)
    row_number: int = Field(ge=1)
    values: tuple[tuple[str, str | None], ...]


class XlsxSheet(ImmutableModel):
    sheet_id: UUID = Field(default_factory=uuid4)
    name: str
    dimension: str
    max_row: int = Field(ge=1)
    max_column: int = Field(ge=1)
    header_row: int | None = Field(default=None, ge=1)
    headers: tuple[str, ...] = ()
    metadata_row_numbers: tuple[int, ...] = ()
    cells: Cells
    records: tuple[XlsxNormalizedRecord, ...] = ()
    warnings: tuple[XlsxWarning, ...] = ()


class XlsxWorkbook(ImmutableModel):
    source_file: SourceFile
    extraction_method: str = "openpyxl.read_only.data_only"
    sheets: Annotated[tuple[XlsxSheet, ...], Field(min_length=1)]
    warnings: tuple[XlsxWarning, ...] = ()


def _source_path(path: Path, dossier_root: Path) -> str:
    try:
        return path.resolve().relative_to(dossier_root.resolve()).as_posix()
    except ValueError as exc:
        raise ValueError(f"workbook path is outside the supplied dossier: {path}") from exc


def _raw_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _normalize_value(value: Any, raw_value: str) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime | date):
        return parse_date(value, locale="de").isoformat(), "date"
    if isinstance(value, bool):
        return "true" if value else "false", "boolean"
    if isinstance(value, int):
        return format(parse_decimal(value, locale="en"), "f"), "decimal"
    if isinstance(value, float):
        return format(parse_decimal(str(value), locale="en"), "f"), "decimal"
    if isinstance(value, Decimal):
        return format(parse_decimal(value, locale="en"), "f"), "decimal"
    if isinstance(value, str):
        if value == "":
            return None, None
        if _GERMAN_DATE.fullmatch(value):
            return parse_date(value, locale="de").isoformat(), "date"
        if _GERMAN_DECIMAL.fullmatch(value):
            return format(parse_decimal(value, locale="de"), "f"), "decimal"
        return normalize_identifier(value, uppercase=False), "identifier"
    return raw_value, "string"


def _formula_cells(path: Path) -> dict[tuple[str, str], str]:
    """Read formulas from package XML while values still come from data-only openpyxl."""

    formulas: dict[tuple[str, str], str] = {}
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            targets = {
                relationship.attrib["Id"]: relationship.attrib["Target"]
                for relationship in relationships
            }
            relationship_attribute = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            for sheet in workbook.iter(
                "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"
            ):
                sheet_name = sheet.attrib["name"]
                target = targets[sheet.attrib[relationship_attribute]]
                archive_path = (
                    target.lstrip("/")
                    if target.startswith("/")
                    else posixpath.normpath(posixpath.join("xl", target))
                )
                sheet_xml = ElementTree.fromstring(archive.read(archive_path))
                namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                for cell in sheet_xml.iter(f"{namespace}c"):
                    formula = cell.find(f"{namespace}f")
                    if formula is not None:
                        formulas[(sheet_name, cell.attrib["r"])] = formula.text or ""
    except (BadZipFile, KeyError, ElementTree.ParseError) as exc:
        raise ValueError(f"invalid XLSX package structure: {path}") from exc
    return formulas


def _sheet_dimensions(sheet: Worksheet) -> tuple[str, int, int]:
    try:
        dimension = sheet.calculate_dimension(force=True)
    except TypeError:
        dimension = sheet.calculate_dimension()
    min_column, min_row, max_column, max_row = range_boundaries(dimension)
    return dimension, max(max_row, min_row, 1), max(max_column, min_column, 1)


def _detect_header(
    *, sheet_name: str, cells_by_row: list[list[XlsxCell]], max_column: int
) -> tuple[int | None, tuple[str, ...], tuple[XlsxWarning, ...]]:
    warnings: list[XlsxWarning] = []
    nonempty_rows = [
        (index + 1, row)
        for index, row in enumerate(cells_by_row)
        if any(cell.raw_value != "" for cell in row)
    ]
    for nonempty_index, (row_number, row) in enumerate(nonempty_rows):
        if nonempty_index == len(nonempty_rows) - 1:
            break
        nonempty = [cell for cell in row if cell.raw_value != ""]
        required = 1 if max_column == 1 else 2
        if len(nonempty) < required or len(nonempty) / max_column < 0.5:
            continue
        if any(cell.value_type not in {"string", "inlineStr", "s"} for cell in nonempty):
            continue

        headers: list[str] = []
        seen: dict[str, int] = {}
        for cell in row:
            if cell.raw_value == "":
                header = f"column_{get_column_letter(cell.column_number)}"
                warnings.append(
                    XlsxWarning(
                        message=f"missing header replaced with {header}",
                        sheet_name=sheet_name,
                        row_number=row_number,
                        cell=cell.coordinate,
                    )
                )
            else:
                header = normalize_identifier(cell.raw_value, uppercase=False)
            occurrence = seen.get(header, 0) + 1
            seen[header] = occurrence
            if occurrence > 1:
                unique_header = f"{header}__{occurrence}"
                warnings.append(
                    XlsxWarning(
                        message=f"duplicate header {header!r} renamed to {unique_header!r}",
                        sheet_name=sheet_name,
                        row_number=row_number,
                        cell=cell.coordinate,
                    )
                )
                header = unique_header
            headers.append(header)
        return row_number, tuple(headers), tuple(warnings)

    warnings.append(
        XlsxWarning(message="no conservative header row detected", sheet_name=sheet_name)
    )
    return None, (), tuple(warnings)


def parse_xlsx_workbook(path: Path, *, dossier_root: Path) -> XlsxWorkbook:
    """Parse all sheets and all cells in their reported dimensions."""

    path = path.expanduser().resolve()
    root = dossier_root.expanduser().resolve()
    source_path = _source_path(path, root)
    source_sha256 = sha256_file(path)
    formulas = _formula_cells(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"could not open XLSX workbook {source_path}: {exc}") from exc

    parsed_sheets: list[XlsxSheet] = []
    workbook_warnings: list[XlsxWarning] = []
    try:
        for sheet in workbook.worksheets:
            dimension, max_row, max_column = _sheet_dimensions(sheet)
            cells_by_row: list[list[XlsxCell]] = []
            sheet_warnings: list[XlsxWarning] = []
            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                ),
                start=1,
            ):
                parsed_row: list[XlsxCell] = []
                for column_number, cell in enumerate(row, start=1):
                    coordinate = f"{get_column_letter(column_number)}{row_number}"
                    value = cell.value
                    raw_value = _raw_string(value)
                    normalized_value, normalization_kind = _normalize_value(value, raw_value)
                    formula = formulas.get((sheet.title, coordinate))
                    extraction_method = (
                        "openpyxl.read_only.data_only.formula_cache"
                        if formula is not None
                        else "openpyxl.read_only.data_only"
                    )
                    if formula is not None and value is None:
                        sheet_warnings.append(
                            XlsxWarning(
                                message="formula has no cached value",
                                sheet_name=sheet.title,
                                row_number=row_number,
                                cell=coordinate,
                            )
                        )
                    value_type = (
                        TYPE_FORMULA
                        if formula is not None
                        else getattr(cell, "data_type", None) or "empty"
                    )
                    evidence = EvidenceRef(
                        source_path=source_path,
                        source_type=SourceType.XLSX_CELL,
                        file_sha256=source_sha256,
                        raw_value=raw_value,
                        normalized_value=normalized_value,
                        extraction_method=extraction_method,
                        row=row_number,
                        sheet=sheet.title,
                        cell=coordinate,
                    )
                    parsed_row.append(
                        XlsxCell(
                            coordinate=coordinate,
                            row_number=row_number,
                            column_number=column_number,
                            raw_value=raw_value,
                            value_type=value_type,
                            normalized_value=normalized_value,
                            normalization_kind=normalization_kind,
                            extraction_method=extraction_method,
                            formula=formula,
                            evidence_ref=evidence,
                        )
                    )
                cells_by_row.append(parsed_row)

            header_row, headers, header_warnings = _detect_header(
                sheet_name=sheet.title,
                cells_by_row=cells_by_row,
                max_column=max_column,
            )
            sheet_warnings.extend(header_warnings)
            metadata_rows = tuple(range(1, header_row)) if header_row is not None else tuple(
                range(1, max_row + 1)
            )
            records = ()
            if header_row is not None:
                records = tuple(
                    XlsxNormalizedRecord(
                        row_number=row_number,
                        values=tuple(
                            (header, cell.normalized_value)
                            for header, cell in zip(
                                headers, cells_by_row[row_number - 1], strict=True
                            )
                        ),
                    )
                    for row_number in range(header_row + 1, max_row + 1)
                )
            parsed_sheets.append(
                XlsxSheet(
                    name=sheet.title,
                    dimension=dimension,
                    max_row=max_row,
                    max_column=max_column,
                    header_row=header_row,
                    headers=headers,
                    metadata_row_numbers=metadata_rows,
                    cells=tuple(cell for row in cells_by_row for cell in row),
                    records=records,
                    warnings=tuple(sheet_warnings),
                )
            )
            workbook_warnings.extend(sheet_warnings)
    finally:
        workbook.close()

    return XlsxWorkbook(
        source_file=SourceFile(
            path=source_path,
            file_type=infer_file_type(path),
            byte_size=path.stat().st_size,
            sha256=source_sha256,
        ),
        sheets=tuple(parsed_sheets),
        warnings=tuple(workbook_warnings),
    )
